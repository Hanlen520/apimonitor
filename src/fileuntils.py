#!/usr/bin/python
# -*- coding: UTF-8 -*-


"""
操作excel工具类
"""


import string
import datetime
from openpyxl import Workbook,load_workbook




def get_es_list():
    """
    获取英文26字母列表
    :return:
    """
    return [e for e in string.ascii_uppercase]


def set_sheet_header(header_name_list,ws):
    """
    设置表格中列的名字
    :param header_name_list:
    :return:
    """
    sheet_es_list = [e for e in string.ascii_uppercase][0:len(header_name_list)]
    for i,h in  zip(sheet_es_list,header_name_list):
        ws['{}1'.format(i)] = h
    return ws


def set_sheet_value(value_name_list,ws):
    """
    设置表格中列的对应的值
    :param value_name_list:
    :param ws:
    :return:
    """

    for v in  value_name_list:
        ws.append(v)
    return ws



def write_excel(sheet_name="sheet",header_list=[],value_list=[],save_excel_name='test.xlsx'):
    """
    创建excel并写入数据
    :param sheet_name:
    :return:
    """
    wb = Workbook()
    ws = wb.create_sheet(sheet_name)
    ws = wb.active
    ws = set_sheet_header(header_list, ws)
    ws = set_sheet_value(value_list, ws)
    wb.save(save_excel_name)


def load_excel(load_excel_name=None,sheet_name='Sheet'):
    """
    读取表格中的所有数据
    :param load_excel_name:
    :param sheet_name:
    :return:
    """
    excel_value_list = []
    wb = load_workbook(load_excel_name)
    sheet = wb.get_sheet_by_name(name=sheet_name)
    for one_column_data in sheet.iter_rows():
        for o in one_column_data:
            if o.value != None:
               value_list.append(o.value)
        excel_value_list.append(value_list)
    return excel_value_list



if __name__ == '__main__':

    sheet_name = 'test'
    header_list = ['序号','姓名']
    value_list = [['mike',18],['tom',20],['lili',23]]
    write_excel(sheet_name=sheet_name,header_list=header_list,value_list=value_list,save_excel_name='persion.xlsx')
    print(load_excel(load_excel_name='persion.xlsx',sheet_name=sheet_name))



